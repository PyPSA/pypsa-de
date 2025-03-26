# -*- coding: utf-8 -*-
"""Input - Output related functions."""  # noqa: A005

import inspect
import json
import logging
import re
import tomllib
from importlib import resources
from os import login_tty
from pathlib import Path
from typing import Callable

import pandas as pd
import pypsa
import yaml
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.worksheet.worksheet import Worksheet
from pandas import ExcelWriter
from pydantic.v1.utils import deep_update
from xlsxwriter.utility import xl_col_to_name, xl_rowcol_to_cell

from evals.configs import ExcelConfig, MetricConfig
from evals.constants import (
    ALIAS_COUNTRY_REV,
    ALIAS_LOCATION,
    ALIAS_REGION_REV,
    RUN_META_DATA,
    DataModel,
    Regex,
)
from evals.utils import (
    calculate_cost_annuity,
    filter_by,
    get_mapping,
    insert_index_level,
    rename_aggregate,
)


def read_networks(result_path: str | Path, sub_directory: str = "networks") -> dict:
    """Read postnetwork results from NetCDF (.nc) files.

    The function returns a dictionary of data frames. The planning
    horizon (year) is used as dictionary key and added to the network
    as an attribute to associate the year with it. Network snapshots
    are equal for all networks, although the year changes. This is
    required to align timestamp columns in a data frame. Snapshots
    will become fixed late in the evaluation process (just before
    export to file).

    In addition, the function patches the statistics accessor attached
    to loaded networks and adds the configuration under n.meta if it is
    missing.

    Parameters
    ----------
    result_path
        Absolute or relative path to the run results folder that
        contains all model results (typically ends with "results",
        or is a time-stamp).
    sub_directory
        The subdirectory name to read files from relative to the
        result folder.

    Returns
    -------
    :
        A Dictionary that contains pypsa.Network objects as values the
        year from the end of the file name as keys.
    """
    # delayed import to prevent circular dependency error
    from statistic import ESMStatistics

    input_path = Path(result_path) / sub_directory
    networks = {}
    for file_path in input_path.glob("*.nc"):
        year = re.search(r"\d{4}$", file_path.stem).group()
        n = pypsa.Network(file_path)
        n.statistics = ESMStatistics(n, result_path)
        networks[year] = n

    assert networks, f"No networks found in {input_path}."

    return networks


#
# def read_model_config(result_path: str | Path, sub_directory: str = "configs") -> dict:
#     """Read the run configuration from a YAML file.
#
#     Parameters
#     ----------
#     result_path
#         Absolute or relative path to the run results folder that
#         contains all model results (typically ends with "results",
#         or is a time-stamp).
#     sub_directory : optional
#         The subdirectory name to read files from relative to the
#         result folder.
#
#     Returns
#     -------
#     :
#         The configuration items in a dictionary.
#     """
#     file_path = Path(result_path) / sub_directory / "config_customize.yaml"
#     with file_path.open("r", encoding="utf-8") as config:
#         return yaml.safe_load(config)
#


def read_views_config(
    func: Callable, config_override: str = "config.override.toml"
) -> dict:
    """Return the configuration for a view function.

    The function reads the default configuration from the
    TOML file and optionally updates it using the config
    file from the override file. The configuration returned
    is stripped down to the relevant parts that matter for the
    called view function.

    Parameters
    ----------
    func
        The view function to be called by the CLI module.
    config_override
        A file name as a string as passed to the CLI module.

    Returns
    -------
    :
        The default configuration with optional overrides from
        a second configuration file.
    """
    func_fp = Path(inspect.getmodule(func).__file__)
    func_name = func_fp.stem
    func_module = func_fp.parent.stem
    defaults_fp = resources.files("evals") / "config.default.toml"
    default = tomllib.load(defaults_fp.open("rb"))
    default_global = default["global"]
    default_view = default["view"][func_module][func_name]

    if config_override:
        override_fp = Path(resources.files("evals")) / config_override
        override = tomllib.load(override_fp.open("rb"))
        default_global = deep_update(default_global, override["global"])

        if (
            override_view := override.get("view", {})
            .get(func_module, {})
            .get(func_name)
        ):
            default_view = deep_update(default_view, override_view)

    config = {"global": default_global, "view": default_view}

    logger = logging.getLogger()
    logger.debug(f"Configuration items: {config}")

    return config


def read_csv_files(
    result_path: str | Path, glob: str, sub_directory: str
) -> pd.DataFrame:
    """Read CSV files from disk.

    Assumes, that if the file name ends with an underscore and 4
    digits, the 4 digits represent the year. The year is prepended
    to the result dataframe index. Otherwise, the first column in the
    CSV file will be the index.

    The function caches result with the same input arguments.

    Parameters
    ----------
    result_path
        Absolute or relative path to the run results folder that
        contains all model results (typically ends with "results",
        or is a time-stamp).
    glob
        The search pattern to filter file names. The asterix can
        be used as wildcard character that matches anything.
    sub_directory
        The subdirectory name to read files from relative to the
        result folder.

    Returns
    -------
    :
        All CSV files concatenated into one DataFrame along the
        index axis.
    """
    input_path = Path(result_path) / sub_directory
    assert input_path.is_dir(), f"Input path does not exist: {input_path.resolve()}"
    file_paths = input_path.glob(glob)

    df_list = []
    for file_path in file_paths:
        _df = pd.read_csv(file_path, index_col=0)
        if year := re.search(Regex.year, file_path.stem):
            _df = insert_index_level(_df, year.group(), DataModel.YEAR)
        df_list.append(_df)

    # must assert after the loop, because file_paths is a generator
    assert df_list, f"No files named like '{glob}' in {input_path.resolve()}."

    return pd.concat(df_list, sort=True)


def prepare_costs(
    result_path: str | Path,
    n_years: int,
    raw: bool = False,
    sub_directory: str = "esm_run/interpolated_data",
) -> pd.DataFrame:
    """Read cost files and calculate fixed costs.

    Parameters
    ----------
    result_path
        Absolute or relative path to the run results folder that
        contains all model results (typically ends with "results",
        or is a time-stamp).
    n_years
        The number of years used to calculate the annual costs.
    raw
        Whether to return the CSV data as is, i.e. without added cost
        calculations.
    sub_directory
        The location of the costs files relative to the results folder
        (Only required during testing).

    Returns
    -------
    :
        Costs by year, technology and parameter.
    """
    costs = read_csv_files(result_path, "costs_*.csv", sub_directory)
    costs = costs.set_index("parameter", append=True, drop=True)
    if raw:
        return costs

    config = read_model_config(result_path)
    default_costs = {
        "CO2 intensity": 0,
        "FOM": 0,
        "VOM": 0,
        "discount rate": config["costs"]["discountrate"],
        "efficiency": 1,
        "fuel": 0,
        "investment": 0,
        "lifetime": config["costs"]["lifetime"],
    }
    costs.loc[costs["unit"].str.contains("/kW"), "value"] *= 1e3  # to 1/MW
    usd_to_eur = config["costs"]["USD2013_to_EUR2013"]
    costs.loc[costs["unit"].str.contains("USD"), "value"] *= usd_to_eur

    costs = costs["value"]
    costs = costs.unstack(level="parameter")
    costs = costs.groupby(costs.index.names).sum(min_count=1)  # keep NaNs
    costs = costs.fillna(default_costs)

    def _calculate_fixed_costs(row: pd.Series) -> pd.Series:
        """Calculate the fixed costs per row.

        Parameters
        ----------
        row
            A row in the costs data frame.

        Returns
        -------
        :
            The fixed costs for the row.
        """
        annuity = calculate_cost_annuity(row["lifetime"], row["discount rate"])
        annuity += row["FOM"] / 100.0
        return annuity * row["investment"] * n_years

    costs["fixed"] = costs.apply(_calculate_fixed_costs, axis=1)

    # operations between data frames are easier if index names align
    costs.index.names = [DataModel.YEAR, DataModel.CARRIER]

    return costs


def prepare_co2_emissions(
    result_path: str | Path,
    options: str | list[str],
    sub_directory: str = "../data",
) -> pd.DataFrame:
    """Read emissions from file and calculate the country share.

    The function reads the emissions from file. It also uses the run
    configuration to determine the cluster settings. If the cluster is
    named 'AT10', the region population shares are read from file.
    The region population shares split the country level emissions into
    region emissions.

    Parameters
    ----------
    result_path
        Absolute or relative path to the run results folder that
        contains all model results (typically ends with "results",
        or is a time-stamp).
    options : {'T', 'H', 'I'}
        Specify 'T', 'H', and/or 'I' to include transport,
        household, and/or industry emissions, respectively.
    sub_directory
        The location of the costs files relative to the results folder
        (Only required during testing).

    Returns
    -------
    :
        The CO2 emission share by country.
    """
    # metric_name = "CO2 Share (1)"
    config = read_model_config(result_path)
    scenario_path = Path(result_path) / sub_directory / config["scenario"]["name"]
    co2 = read_csv_files(scenario_path, "co2_totals.csv", "general")
    co2 = co2.drop("EU28")  # The EU value distorts the sum

    cols = ["electricity"]
    if "T" in options:  # Transport
        cols.extend(["rail non-elec", "road non-elec"])
    if "H" in options:  # Households
        cols.extend(["residential non-elec", "services non-elec"])
    if "I" in options:  # Industry
        cols.extend(
            [
                "industrial non-elec",
                "industrial processes",
                "domestic aviation",
                "domestic navigation",
            ]
        )

    co2 = co2[cols].sum(axis=1)
    co2_dist = co2 / co2.sum()
    cluster = config["scenario"]["clusters"][0]
    if cluster != "AT10":
        return co2_dist  # .to_frame(metric_name)

    # disaggregate AT cluster emissions by population share
    pop_at = read_csv_files(scenario_path, "custom_urban_pop_share_AT10.csv", cluster)
    pop_dist_at = pop_at["region_pop"] / pop_at["region_pop"].sum()
    co2_dist_at = pop_dist_at * co2_dist["AT"]

    co2_dist = pd.concat([co2_dist, co2_dist_at], axis=0, sort=True)
    co2_dist = co2_dist.drop("AT")
    co2_dist.index = co2_dist.index.set_names(DataModel.LOCATION)

    return co2_dist


def prepare_industry_demand(
    result_path: str | Path,
    networks: dict,
    sub_directory: str = "interpolated_data",
    rename: bool = True,
) -> pd.Series:
    """Read industry demand from resource file and correct methane.

    The industry demand for methane must be increased by the carbon
    capture amounts. The 'gas for industry CC' Links have efficiencies
    smaller than 1. Therefore, resulting losses must be added to the
    methane values in the resource file.

    The industry demand is read from resource files, and not
    extracted from the postnetwork, because the postnetwork contains
    demands in an aggregated form and the CSV files are on a per-sector
    granularity.

    Parameters
    ----------
    result_path
        Absolute or relative path to the run results folder that
        contains all model results. (typically ends with "results",
        or is a time-stamp).
    networks
        The pypsa postnetworks, used to extract the methane amounts
        including CC.
    sub_directory : optional
        The location of the CSV files relative to the results folder.
    rename : optional
        Whether, or not to rename sectors and apply aggregate groups.

    Returns
    -------
    :
        The industry demand by year, country, and sector group in MWh.
    """
    # delayed import to prevent circular import error
    from statistic import collect_myopic_statistics

    config = read_model_config(result_path)
    simpl = config["scenario"]["simpl"][0]
    cluster = config["scenario"]["clusters"][0]

    file_name_pattern = f"industrial_demand_by_sector_elec_s{simpl}_{cluster}_*.csv"
    industry_demand = read_csv_files(result_path, file_name_pattern, sub_directory)
    industry_demand = industry_demand.set_index("sector", append=True)

    if rename:
        industry_sectors = {
            "Integrated steelworks": "Iron and Steel",
            "Electric arc": "Iron and Steel",
            "Cement": "Non-metallic Minerals",
            "Ceramics & other NMM": "Non-metallic Minerals",
            "Glass production": "Non-metallic Minerals",
            "Basic chemicals": "Chemical",
            "Other chemicals": "Chemical",
            "Pharmaceutical products etc.": "Chemical",
            "Pulp production": "Paper, Pulp and Print",
            "Paper production": "Paper, Pulp and Print",
            "Printing and media reproduction": "Paper, Pulp and Print",
            "Alumina production": "Non-ferrous Metals",
            "Aluminium - primary production": "Non-ferrous Metals",
            "Aluminium - secondary production": "Non-ferrous Metals",
            "Other non-ferrous metals": "Non-ferrous Metals",
            "Food, beverages and tobacco": "Food and Tabacco",
            "Transport Equipment": "Transport Equipment",
            "Machinery Equipment": "Machinery",
            "Textiles and leather": "Textile and Leather",
            "Wood and wood products": "Wood and Wood Products",
            "Other Industrial Sectors": "Non-specified Industry",
        }
        industry_demand = rename_aggregate(
            industry_demand, industry_sectors, level="sector"
        )

    industry_demand = industry_demand * 1e6  # to MWh base unit

    gas_for_industry = collect_myopic_statistics(
        networks,
        "withdrawal",
        comps="Link",
        bus_carrier="gas",
        carrier=["gas for industry", "gas for industry CC"],
        drop_zero_rows=False,
        at_port=False,  # only port 0
    )

    # combine gas demand (CC and non-CC) by naming them the same
    mapper = {"gas for industry CC": "gas for industry"}
    gas_for_industry = rename_aggregate(
        gas_for_industry, mapper, level=DataModel.CARRIER
    )

    def increase_gas_demand_by_cc_share(ser: pd.Series) -> pd.Series:
        """Correct the gas demand for industry by CC fraction.

        The value in the resource file does not include energy for
        carbon capture. The sector values must be increased by
        the amount of energy used by carbon capture (if any).

        Parameters
        ----------
        ser
            The input Series with gas demand values.

        Returns
        -------
        :
            Gas demands increased by the energy amounts needed for
            carbon capture.
        """
        demand_cc = gas_for_industry[pd.IndexSlice[ser.name]].to_numpy()[0]
        demand = ser.sum()
        scaling_factor = demand_cc / demand
        return ser if demand < 1 else ser * scaling_factor

    industry_demand["methane"] = (
        industry_demand["methane"]
        .groupby(["year", "country"], group_keys=False)
        .apply(increase_gas_demand_by_cc_share)
    )

    # rename index levels for consistency with the statistics data model
    # The 'sector' in the industry demand interpolated data file becomes
    # the DataModel 'bus_carrier'. The unnamed column (technologies?)
    # becomes the DataModel 'carrier'
    industry_demand = industry_demand.stack().swaplevel(-1, -2)
    industry_demand.index.names = [DataModel.YEAR] + DataModel.IDX_NAMES

    return industry_demand


def prepare_nodal_energy(
    result_path: str | Path, sub_directory: str = "../resources"
) -> pd.DataFrame:
    """Prepare nodal energy data for analysis.

    Since bus_carrier information is not included in the nodal energy
    CSV files, it is explicitly added here through a mapping of CSV
    file column labels to bus_carrier names.

    Note that all values returned are negative, although demand usually
    has negative sign in the evaluation routines.

    Parameters
    ----------
    result_path
        Absolute or relative path to the results folder in the project
        root directory.
    sub_directory
        The subdirectory relative to the result path where the CSV
        files are located.

    Returns
    -------
    :
        Processed nodal energy data with aggregated locations and
        scaled to MWh.

    Notes
    -----
    The result does not contain bus_carrier information. This should
    be added here
    """
    statistic_name = "Energy Totals"
    statistic_unit = "MWh"
    try:
        nodal_energy = read_csv_files(
            result_path, "nodal_energy_totals_*.csv", sub_directory=sub_directory
        )  # in TWh
    except AssertionError:
        nodal_energy = read_csv_files(
            result_path,
            "nodal_energy_totals_*.csv",
            sub_directory=sub_directory.replace("../", "./"),
        )  # in TWh

    nodal_energy = nodal_energy.stack()
    nodal_energy.index.names = DataModel.YEAR_IDX_NAMES[:3]
    nodal_energy = nodal_energy * 1e6  # hard cast to MWh

    carrier_bus_carrier_map = {
        "BEV road freight LNF": "AC",
        "BEV road freight Lkw<12t": "AC",
        "BEV road freight Lkw>12t": "AC",
        "BEV road passenger": "AC",
        "Benzin HEV road passenger": "Fischer-Tropsch",
        "Benzin PHEV road passenger": "Fischer-Tropsch",
        "Benzin road passenger": "Fischer-Tropsch",
        "CNG HEV road passenger": "gas",
        "CNG PHEV road passenger": "gas",
        "CNG road freight LNF": "gas",
        "CNG road freight Lkw<12t": "gas",
        "CNG road freight Lkw>12t": "gas",
        "CNG road passenger": "gas",
        "Diesel HEV road passenger": "Fischer-Tropsch",
        "Diesel PHEV road freight LNF": "Fischer-Tropsch",
        "Diesel PHEV road passenger": "Fischer-Tropsch",
        "Diesel road freight LNF": "Fischer-Tropsch",
        "Diesel road freight Lkw<12t": "Fischer-Tropsch",
        "Diesel road freight Lkw>12t": "Fischer-Tropsch",
        "Diesel road passenger": "Fischer-Tropsch",
        "Diesel-electric hybrid road freight Lkw>12t": "Fischer-Tropsch",
        "Fischer-Tropsch Extra-EU aviation passenger": "Fischer-Tropsch",
        "Fischer-Tropsch Intra-EU aviation passenger": "Fischer-Tropsch",
        "Fischer-Tropsch domestic aviation": "Fischer-Tropsch",
        "Fischer-Tropsch domestic aviation freight": "Fischer-Tropsch",
        "Fischer-Tropsch domestic aviation passenger": "Fischer-Tropsch",
        "Fischer-Tropsch domestic navigation": "Fischer-Tropsch",
        "Fischer-Tropsch domestic navigation freight": "Fischer-Tropsch",
        "Fischer-Tropsch domestic navigation passenger": "Fischer-Tropsch",
        "Fischer-Tropsch heavy duty road freight": "Fischer-Tropsch",
        "Fischer-Tropsch international aviation": "Fischer-Tropsch",
        "Fischer-Tropsch international aviation freight": "Fischer-Tropsch",
        "Fischer-Tropsch international aviation passenger": "Fischer-Tropsch",
        "Fischer-Tropsch international navigation": "Fischer-Tropsch",
        "Fischer-Tropsch international navigation freight": "Fischer-Tropsch",
        "Fischer-Tropsch international navigation passenger": "Fischer-Tropsch",
        "Fischer-Tropsch light duty road freight": "Fischer-Tropsch",
        "Fischer-Tropsch other road passenger": "Fischer-Tropsch",
        "Fischer-Tropsch passenger cars": "Fischer-Tropsch",
        "Fischer-Tropsch rail": "Fischer-Tropsch",
        "Fischer-Tropsch rail freight": "Fischer-Tropsch",
        "Fischer-Tropsch rail passenger": "Fischer-Tropsch",
        "Fischer-Tropsch road": "Fischer-Tropsch",
        "Fischer-Tropsch road freight": "Fischer-Tropsch",
        "Fischer-Tropsch road passenger": "Fischer-Tropsch",
        "Fischer-Tropsch two-wheel": "Fischer-Tropsch",
        "H2 FCV road freight LNF": "H2",
        "H2 FCV road freight Lkw<12t": "H2",
        "H2 FCV road freight Lkw>12t": "H2",
        "H2 FCV road passenger": "H2",
        "H2 domestic aviation": "H2",
        "H2 domestic navigation": "H2",
        "H2 domestic navigation freight": "H2",
        "H2 domestic navigation passenger": "H2",
        "H2 international aviation": "H2",
        "H2 international navigation": "H2",
        "H2 international navigation freight": "H2",
        "H2 international navigation passenger": "H2",
        "H2 rail": "H2",
        "H2 rail passenger": "H2",
        "H2 road": "H2",
        "H2 road freight": "H2",
        "H2 road freight LNF": "H2",
        "H2 road freight Lkw<12t": "H2",
        "H2 road freight Lkw>12t": "H2",
        "H2 road passenger": "H2",
        "LH2 Extra-EU aviation passenger": "LH2",
        "LH2 Intra-EU aviation passenger": "LH2",
        "LH2 domestic aviation freight": "LH2",
        "LH2 domestic aviation passenger": "LH2",
        "LH2 domestic navigation freight": "LH2",
        "LH2 domestic navigation passenger": "LH2",
        "LH2 international aviation freight": "LH2",
        "LH2 international aviation passenger": "LH2",
        "LH2 international navigation freight": "LH2",
        "LH2 international navigation passenger": "LH2",
        "LH2 rail freight": "LH2",
        "LNG domestic navigation freight": "gas",
        "LNG international navigation freight": "gas",
        "LNG road freight Lkw<12t": "gas",
        "LNG road freight Lkw>12t": "gas",
        "biogas light duty road freight": "gas",
        "biogas other road passenger": "gas",
        "biogas passenger cars": "gas",
        "biogas road": "gas",
        "electricity light duty road freight": "AC",
        "electricity other road passenger": "AC",
        "electricity passenger cars": "AC",
        "electricity rail": "AC",
        "electricity rail freight": "AC",
        "electricity rail passenger": "AC",
        "electricity residential": "AC",
        "electricity residential cooking": "AC",
        "electricity residential space": "AC",
        "electricity residential water": "AC",
        "electricity road": "AC",
        "electricity road freight": "AC",
        "electricity road passenger": "AC",
        "electricity services": "AC",
        "electricity services cooking": "AC",
        "electricity services space": "AC",
        "electricity services water": "AC",
        "gas domestic navigation": "gas",
        "gas domestic navigation passenger": "gas",
        "gas international navigation": "gas",
        "gas international navigation passenger": "gas",
        "gas light duty road freight": "gas",
        "gas other road passenger": "gas",
        "gas passenger cars": "gas",
        "gas road": "gas",
        "gas road freight": "gas",
        "gas road passenger": "gas",
        "oil rail": "Fischer-Tropsch",
        "oil road": "Fischer-Tropsch",
        "renewables rail": "AC",
        "renewables road": "AC",
        "total aviation freight": "mixed",
        "total aviation passenger": "mixed",
        "total domestic aviation": "mixed",
        "total domestic aviation freight": "mixed",
        "total domestic aviation passenger": "mixed",
        "total domestic navigation": "mixed",
        "total electricity equivalent passenger": "mixed",
        "total heavy duty road freight": "mixed",
        "total international aviation": "mixed",
        "total international aviation freight": "mixed",
        "total international aviation passenger": "mixed",
        "total international navigation": "mixed",
        "total light duty road freight": "mixed",
        "total other road passenger": "mixed",
        "total passenger cars": "mixed",
        "total rail": "mixed",
        "total rail freight": "mixed",
        "total rail passenger": "mixed",
        "total residential": "mixed",
        "total residential cooking": "mixed",
        "total residential space": "mixed",
        "total residential water": "mixed",
        "total road": "mixed",
        "total road freight": "mixed",
        "total road passenger": "mixed",
        "total services": "mixed",
        "total services cooking": "mixed",
        "total services space": "mixed",
        "total services water": "mixed",
        "total two-wheel": "mixed",
    }

    carrier = nodal_energy.index.get_level_values("carrier")
    nodal_energy = nodal_energy.to_frame()
    nodal_energy["bus_carrier"] = carrier.map(carrier_bus_carrier_map)
    nodal_energy = nodal_energy.set_index("bus_carrier", append=True, drop=True)
    nodal_energy = nodal_energy.squeeze()

    nodal_energy.attrs["name"] = statistic_name
    nodal_energy.attrs["unit"] = statistic_unit

    return nodal_energy


def export_excel_countries(
    metric: pd.DataFrame,
    writer: pd.ExcelWriter,
    cfg: MetricConfig,
) -> None:
    """Add one sheet per country to an Excel file.

    The function appends one sheet per location to the workbook of the
    opened writer instance.

    Parameters
    ----------
    metric
        The data frame without carrier mapping applied.
    writer
        The ExcelWriter instance to add the sheets to.
    cfg
        The Excel file configuration.
    """
    mapping_int = get_mapping(cfg.mapping, "internal")
    mapping_ext = get_mapping(cfg.mapping, "external")

    carrier = metric.index.unique(DataModel.CARRIER)
    df = rename_aggregate(metric, level=DataModel.CARRIER, mapper=mapping_int)
    df = filter_by(df, location=list(ALIAS_COUNTRY_REV))  # exclude regions
    df = df.pivot_table(
        index=cfg.excel.pivot_index, columns=cfg.excel.pivot_columns, aggfunc="sum"
    )

    for country, data in df.groupby(DataModel.LOCATION):
        data = data.droplevel(DataModel.LOCATION)
        _write_excel_sheet(data, cfg.excel, writer, str(country))

    _write_mapping_sheet(mapping_int, carrier, writer, sheet_name="Internal Mapping")
    _write_mapping_sheet(mapping_ext, carrier, writer, sheet_name="External Mapping")


def export_excel_regions_at(
    metric: pd.DataFrame,
    writer: pd.ExcelWriter,
    cfg: MetricConfig,
) -> None:
    """Write one Excel sheet for Europe, Austria, and Austrian regions.

    The function appends one sheet per location to the workbook of the
    opened writer instance.

    Parameters
    ----------
    metric
        The data frame without carrier mapping applied.
    writer
        The ExcelWriter instance to add the sheets to.
    cfg
        The Excel file configuration.
    """
    mapping_int = get_mapping(cfg.mapping, "internal")
    mapping_ext = get_mapping(cfg.mapping, "external")

    carrier = metric.index.unique(DataModel.CARRIER)
    df = rename_aggregate(metric, level=DataModel.CARRIER, mapper=mapping_int)
    df = filter_by(df, location=list(ALIAS_REGION_REV))
    df_xlsx = df.pivot_table(
        index=cfg.excel.pivot_index, columns=cfg.excel.pivot_columns, aggfunc="sum"
    )

    for country, data in df_xlsx.groupby(DataModel.LOCATION):
        data = data.droplevel(DataModel.LOCATION)
        _write_excel_sheet(data, cfg.excel, writer, str(country))

    # append carrier tables to special region sheet
    df_region = df.pivot_table(
        index=DataModel.CARRIER,
        columns=[DataModel.LOCATION, DataModel.YEAR],
        aggfunc="sum",
    ).droplevel(DataModel.METRIC, axis=1)

    cfg.excel.chart_title = "Region AT"
    _write_excel_sheet(
        df_region,
        cfg.excel,
        writer,
        sheet_name="Regions AT",
        position=3,
    )
    groups = df_region.drop(
        ["Europe", "Austria"], level=DataModel.LOCATION, axis=1, errors="ignore"
    ).groupby(DataModel.CARRIER)

    # update config for pivoted carrier tables and graphs
    cfg.chart = "clustered"
    cfg.chart_switch_axis = True

    for carrier, df_reg in groups:
        cfg.chart_title = str(carrier).title()
        _write_excel_sheet(
            df_reg.T.unstack(1),
            cfg.excel,
            writer,
            sheet_name="Regions AT",
            position=3,
        )

    _write_mapping_sheet(mapping_int, carrier, writer, sheet_name="Internal Mapping")
    _write_mapping_sheet(mapping_ext, carrier, writer, sheet_name="External Mapping")


def _write_excel_sheet(
    df: pd.DataFrame,
    cfg: ExcelConfig,
    writer: pd.ExcelWriter,
    sheet_name: str,
    position: int = -1,
) -> None:
    """Write a data frame to an Excel sheet.

    The input data are written to xlsx and a corresponding diagram
    (currently only stacked bar chart) is included.

    Parameters
    ----------
    df
        The dataframe to be transformed and exported to Excel; works
        with columns of multiindex level <= 2 f.ex. (location, year).
    cfg
        The configuration of the Excel file and chart.
    writer
        The writer object that represents an opened Excel file.
    sheet_name
        The name of sheet included in xlsx, will also be the
        name of diagram.
    position
        The position where the worksheet should
        be added.
    """
    axis_labels = cfg.axis_labels or [df.attrs["name"], df.attrs["unit"]]

    # parametrize size of data in xlsx
    number_rows, number_col = df.shape

    start_row = 0
    if ws := writer.sheets.get(sheet_name):
        # the sheet already exists. We can determine the
        # number of rows contained and append new data below
        gap_size = 2 if ws.max_row > 0 else 0
        start_row = ws.max_row + gap_size

    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, float_format="%0.4f")
    ws = writer.sheets.get(sheet_name)  # needed to update ws object

    _delete_index_name_row(ws, df, start_row=start_row)
    _expand_column_to_fit_content(ws, df, 0)

    if cfg.chart:
        barchart = _create_excel_barchart(ws, df, cfg, axis_labels, start_row)
        chart_start_cell = xl_rowcol_to_cell(start_row, number_col + 2)
        ws.add_chart(barchart, chart_start_cell)

    _move_excel_sheet(writer, sheet_name, position)


def _write_mapping_sheet(
    mapping: dict, carrier: tuple, writer: ExcelWriter, sheet_name: str
) -> None:
    """Write the mapping to a separate Excel sheet.

    This is useful to make the renaming process transparent. The mapping
    sheet will show 2 columns: one for the model names and the other for
    the group names (the names also visible in HTML figures).

    Parameters
    ----------
    mapping
        The model name (bus carrier, carrier, or sector) to group
        relation as key value pairs.
    carrier
        A collection of all carrier technologies present in the
        exported metric.
    writer
        The open ExcelWriter object.
    sheet_name
        The name of the sheet to write the 2 mapping columns to.
    """
    m = {k: v for k, v in mapping.items() if k in carrier}
    df = pd.DataFrame.from_dict(m, orient="index", columns=["Alias"])
    df.columns.name = "Carrier"
    df.to_excel(writer, sheet_name=sheet_name, float_format="%0.4f")
    ws = writer.sheets.get(sheet_name)
    _delete_index_name_row(ws, df, start_row=0)  # delete index name row
    _expand_column_to_fit_content(ws, df, 0)
    _expand_column_to_fit_content(ws, df, 1)


def _delete_index_name_row(ws: Worksheet, df: pd.DataFrame, start_row: int) -> None:
    """Remove the index name row from the Excel sheet.

    Delete the row in the Excel worksheet based on the number of index
    levels and the starting row.

    Parameters
    ----------
    ws
        The worksheet where the row will be deleted.
    df
        The DataFrame used to determine the number of index levels.
    start_row
        The starting row from which deletion will begin.
    """
    ws.delete_rows(df.columns.nlevels + 1 + start_row)


def _move_excel_sheet(writer: ExcelWriter, sheet_name: str, position: int) -> None:
    """Move an Excel sheet to a given position.

    Parameters
    ----------
    writer
        The writer instance that depicts an open Excel workbook.
    sheet_name
        The name of the sheet inside the workbook.
    position
        The wanted position of the sheet as integer (1-indexed).
        String input is kept for backwards compatibility and should
        not be used.

    Returns
    -------
    :
        Moves the work sheet to the requested position.
    """
    if position != -1:
        wb = writer.book
        offset = position - len(wb.sheetnames)
        wb.move_sheet(writer.sheets[sheet_name], offset=offset)


def _create_excel_barchart(
    ws: Worksheet,
    df: pd.DataFrame,
    cfg: ExcelConfig,
    axis_labels: list,
    start_row: int,
) -> BarChart:
    """Create an Excel bar chart object.

    The function support bar chart of two different orientations:
      - categories from column labels, and
      - categories from index labels

    The bar chart position is to the right of the newly added data
    that serves as a data references for the chart.

    Parameters
    ----------
    ws
        The open worksheet instance.
    df
        Reference data for the bar chart.
    cfg
        The configuration for the Excel file and Excel chart.
    axis_labels
        A list of strings. The first list item is the x-axis label,
        the second is the y-axis label.
    start_row
        First row for data insertion.

    Returns
    -------
    :
        The Excel bar chart object ready for insertion in a sheet.
    """
    nrows, ncols = df.shape

    # gapWidth controls the space between outer column level groups
    cat_len = len(df.columns.unique(1))
    chart_kwargs = {"gapWidth": 20} if cat_len > 1 else {}
    chart = BarChart(**chart_kwargs)

    # data includes index names left to numeric data
    min_col = df.index.nlevels
    min_row = df.columns.nlevels + 1 + start_row
    max_col = df.index.nlevels + ncols  # same
    max_row = df.columns.nlevels + start_row + nrows  # same

    if cfg.chart_switch_axis:
        # use column names as x-axis labels: the upper left cell shifts
        min_col += 1  # one to the right
        min_row -= 1  # one up

    data = Reference(ws, min_col, min_row, max_col, max_row)

    # reference is the horizontal header innermost one or two rows,
    # or the index in case switch_row_col is True
    min_col = df.index.nlevels + 1
    max_col = df.index.nlevels + ncols
    min_row = start_row + df.columns.nlevels - (1 if cat_len > 1 else 0)
    max_row = start_row + df.columns.nlevels

    if cfg.chart_switch_axis:
        # the categories become the index names of the column names
        # instead. The reference area transposes from row selection to
        # column selection.
        min_col = max_col = min_col - 1  # one left
        min_row += 2  # one down
        max_row = df.columns.nlevels + start_row + nrows

    cats = Reference(ws, min_col, min_row, max_col, max_row)

    chart.type = "col"
    chart.grouping = cfg.chart
    title = cfg.chart_title or df.columns.get_level_values("metric")[0]
    chart.title = title.format(location=ws.title, unit=axis_labels[1])

    # only stack chart if stacked otherwise 0 to avoid always
    # plotting a stacked chart (=100%)
    chart.overlap = 100 if cfg.chart == "stacked" else 0

    chart.add_data(data, from_rows=not cfg.chart_switch_axis, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.title = axis_labels[0]
    chart.y_axis.title = axis_labels[1]
    # excel row height is in pixel. resolution is 72 DPI.
    # default row height in Excel is 15
    height_factor = 0.53  # 15 (px) / 72 (px/Inch) * 2.54 (cm/Inch) = (cm)
    height_data = nrows + df.columns.nlevels
    chart.height = min(height_data * height_factor, 10)
    chart.width = cfg.chart_width  # cm

    if cfg.chart_switch_axis:  # the legend is redundant, as is the x-axis label
        chart.legend = None
        chart.x_axis.title = None

    # set bar colors in chart
    for i, carrier in enumerate(df.index):
        if not cfg.chart_switch_axis:
            color = cfg.chart_colors.get(carrier)
        else:
            color = cfg.chart_colors.get(df.columns.unique("carrier")[0])

        if not color:
            continue

        if not cfg.chart_switch_axis:
            chart.series[i].graphicalProperties.solidFill = color
            chart.series[i].graphicalProperties.line.solidFill = color
        else:
            for ser in chart.series:
                single_bar = DataPoint(idx=i)
                single_bar.graphicalProperties.solidFill = color
                # white borders to separate neighbors of same color
                single_bar.graphicalProperties.line.solidFill = "FFFFFF"
                ser.data_points.append(single_bar)

    return chart


def _expand_column_to_fit_content(ws: Worksheet, df: pd.DataFrame, col: int) -> None:
    """Expand cell columns to improve readability in Excel.

    The function expands the column to the larger value of its current
    column width and the largest string in the input data frame index.

    Parameters
    ----------
    ws
        The open work sheet instance.
    df
        The data added to the worksheet.
    col
        The index of the column that should become expanded.
    """
    xl_col = xl_col_to_name(col)
    series = df.index if col == 0 else df[df.columns[col - 1]]
    data_width = series.astype(str).str.len().max()
    existing_width = ws.column_dimensions[xl_col].width
    column_width = max(existing_width, data_width)
    ws.column_dimensions[xl_col].width = column_width


def export_vamos_jsons(json_file_paths: list, file_name_template: str) -> None:
    """Write a JSON file for VAMOS UI that lists available figures.

    Parameters
    ----------
    json_file_paths
        A collection of file paths with JSON encoded plotly figures.
    file_name_template
        The file name template for the JSON files with unsubstituted
        variables.
    """
    directories = {fp.parent for fp in json_file_paths}
    assert len(directories) == 1, f"Multiple directories are not allowed: {directories}"

    json_file_names = [fp.stem for fp in json_file_paths]

    template_field_values = {}
    if "{year}" in file_name_template:
        template_field_values["Years"] = sorted(
            {re.search(r"_\d{4}_", s).group().strip("_") for s in json_file_names}
        )
    if "{location}" in file_name_template:
        template_field_values["Regions"] = [
            loc
            for loc in ALIAS_LOCATION
            if any(fn.endswith(loc) for fn in json_file_names)
        ]

    payload = [
        {"naming_convention": f"{file_name_template}.json"},
        template_field_values,
        ALIAS_LOCATION,
        RUN_META_DATA,
    ]

    directory = directories.pop()
    file_name = file_name_template.split("_{")[0]
    file_path = directory / f"{file_name}_FILTERS.json"
    with file_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.worksheet.worksheet import Worksheet
from pandas import ExcelWriter
from xlsxwriter.utility import xl_col_to_name, xl_rowcol_to_cell

from evals.configs import ExcelConfig
from evals.constants import ALIAS_COUNTRY_REV, ALIAS_REGION_REV, DataModel
from evals.utils import filter_by, rename_aggregate


def export_excel_countries(
    metric: pd.DataFrame,
    writer: pd.ExcelWriter,
    excel_defaults: ExcelConfig,
    view_config: dict,
) -> None:
    """
    Add one sheet per country to an Excel file.

    The function appends one sheet per location to the workbook of the
    opened writer instance.

    Parameters
    ----------
    metric
        The data frame without carrier mapping applied.
    writer
        The ExcelWriter instance to add the sheets to.
    excel_defaults
        The default settings for Excel file export.
    view_config
        The view configuration items.
    """
    categories = view_config["categories"]
    carrier = metric.index.unique(DataModel.CARRIER)
    df = rename_aggregate(metric, level=DataModel.CARRIER, mapper=categories)
    df = filter_by(df, location=list(ALIAS_COUNTRY_REV))  # exclude regions
    df = df.pivot_table(
        index=excel_defaults.pivot_index,
        columns=excel_defaults.pivot_columns,
        aggfunc="sum",
    )

    for country, data in df.groupby(DataModel.LOCATION):
        data = data.droplevel(DataModel.LOCATION)
        _write_excel_sheet(data, excel_defaults, writer, str(country))

    _write_categories_sheet(categories, carrier, writer, sheet_name="Categories")


def export_excel_regions_at(
    metric: pd.DataFrame,
    writer: pd.ExcelWriter,
    excel_defaults: ExcelConfig,
    view_config: dict,
) -> None:
    """
    Write one Excel sheet for Europe, Austria, and Austrian regions.

    The function appends one sheet per location to the workbook of the
    opened writer instance.

    Parameters
    ----------
    metric
        The data frame without carrier mapping applied.
    writer
        The ExcelWriter instance to add the sheets to.
    excel_defaults
        The default settings for Excel file export.
    view_config
        The view configuration items.
    """
    categories = view_config["categories"]
    carrier = metric.index.unique(DataModel.CARRIER)
    df = rename_aggregate(metric, level=DataModel.CARRIER, mapper=categories)
    df = filter_by(df, location=list(ALIAS_REGION_REV))
    df_xlsx = df.pivot_table(
        index=excel_defaults.pivot_index,
        columns=excel_defaults.pivot_columns,
        aggfunc="sum",
    )

    for country, data in df_xlsx.groupby(DataModel.LOCATION):
        data = data.droplevel(DataModel.LOCATION)
        _write_excel_sheet(data, excel_defaults, writer, str(country))

    # append carrier tables to special region sheet
    df_region = df.pivot_table(
        index=DataModel.CARRIER,
        columns=[DataModel.LOCATION, DataModel.YEAR],
        aggfunc="sum",
    ).droplevel(DataModel.METRIC, axis=1)

    excel_defaults.chart_title = "Region AT"
    _write_excel_sheet(
        df_region,
        excel_defaults,
        writer,
        sheet_name="Regions AT",
        position=3,
    )
    groups = df_region.drop(
        ["Europe", "Austria"], level=DataModel.LOCATION, axis=1, errors="ignore"
    ).groupby(DataModel.CARRIER)

    # update config for pivoted carrier tables and graphs
    excel_defaults.chart = "clustered"
    excel_defaults.chart_switch_axis = True

    for carrier, df_reg in groups:
        excel_defaults.chart_title = str(carrier).title()
        _write_excel_sheet(
            df_reg.T.unstack(1),
            excel_defaults,
            writer,
            sheet_name="Regions AT",
            position=3,
        )

    _write_categories_sheet(categories, carrier, writer, sheet_name="Categories")


def _write_excel_sheet(
    df: pd.DataFrame,
    excel_defaults: ExcelConfig,
    writer: pd.ExcelWriter,
    sheet_name: str,
    position: int = -1,
) -> None:
    """
    Write a data frame to an Excel sheet.

    The input data are written to xlsx and a corresponding diagram
    (currently only stacked bar chart) is included.

    Parameters
    ----------
    df
        The dataframe to be transformed and exported to Excel; works
        with columns of multiindex level <= 2 f.ex. (location, year).
    excel_defaults
        The configuration of the Excel file and chart.
    writer
        The writer object that represents an opened Excel file.
    sheet_name
        The name of sheet included in xlsx, will also be the
        name of diagram.
    position
        The position where the worksheet should
        be added.
    """
    axis_labels = excel_defaults.axis_labels or [df.attrs["name"], df.attrs["unit"]]

    # parametrize size of data in xlsx
    number_rows, number_col = df.shape

    start_row = 0
    if ws := writer.sheets.get(sheet_name):
        # the sheet already exists. We can determine the
        # number of rows contained and append new data below
        gap_size = 2 if ws.max_row > 0 else 0
        start_row = ws.max_row + gap_size

    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, float_format="%0.4f")
    ws = writer.sheets.get(sheet_name)  # needed to update ws object

    _delete_index_name_row(ws, df, start_row=start_row)
    _expand_column_to_fit_content(ws, df, 0)

    if excel_defaults.chart:
        barchart = _create_excel_barchart(
            ws, df, excel_defaults, axis_labels, start_row
        )
        chart_start_cell = xl_rowcol_to_cell(start_row, number_col + 2)
        ws.add_chart(barchart, chart_start_cell)

    _move_excel_sheet(writer, sheet_name, position)


def _write_categories_sheet(
    mapping: dict, carrier: tuple, writer: ExcelWriter, sheet_name: str
) -> None:
    """
    Write the mapping to a separate Excel sheet.

    This is useful to make the renaming process transparent. The mapping
    sheet will show 2 columns: one for the model names and the other for
    the group names (the names also visible in HTML figures).

    Parameters
    ----------
    mapping
        The model name (bus carrier, carrier, or sector) to group
        relation as key value pairs.
    carrier
        A collection of all carrier technologies present in the
        exported metric.
    writer
        The open ExcelWriter object.
    sheet_name
        The name of the sheet to write the 2 mapping columns to.
    """
    m = {k: v for k, v in mapping.items() if k in carrier}
    df = pd.DataFrame.from_dict(m, orient="index", columns=["Category"])
    df.columns.name = "Carrier"
    df.to_excel(writer, sheet_name=sheet_name, float_format="%0.4f")
    ws = writer.sheets.get(sheet_name)
    _delete_index_name_row(ws, df, start_row=0)  # delete index name row
    _expand_column_to_fit_content(ws, df, 0)
    _expand_column_to_fit_content(ws, df, 1)


def _delete_index_name_row(ws: Worksheet, df: pd.DataFrame, start_row: int) -> None:
    """
    Remove the index name row from the Excel sheet.

    Delete the row in the Excel worksheet based on the number of index
    levels and the starting row.

    Parameters
    ----------
    ws
        The worksheet where the row will be deleted.
    df
        The DataFrame used to determine the number of index levels.
    start_row
        The starting row from which deletion will begin.
    """
    ws.delete_rows(df.columns.nlevels + 1 + start_row)


def _move_excel_sheet(writer: ExcelWriter, sheet_name: str, position: int) -> None:
    """
    Move an Excel sheet to a given position.

    Parameters
    ----------
    writer
        The writer instance that depicts an open Excel workbook.
    sheet_name
        The name of the sheet inside the workbook.
    position
        The wanted position of the sheet as integer (1-indexed).
        String input is kept for backwards compatibility and should
        not be used.

    Returns
    -------
    :
        Moves the work sheet to the requested position.
    """
    if position != -1:
        wb = writer.book
        offset = position - len(wb.sheetnames)
        wb.move_sheet(writer.sheets[sheet_name], offset=offset)


def _create_excel_barchart(
    ws: Worksheet,
    df: pd.DataFrame,
    cfg: ExcelConfig,
    axis_labels: list,
    start_row: int,
) -> BarChart:
    """
    Create an Excel bar chart object.

    The function support bar chart of two different orientations:
      - categories from column labels, and
      - categories from index labels

    The bar chart position is to the right of the newly added data
    that serves as a data references for the chart.

    Parameters
    ----------
    ws
        The open worksheet instance.
    df
        Reference data for the bar chart.
    cfg
        The configuration for the Excel file and Excel chart.
    axis_labels
        A list of strings. The first list item is the x-axis label,
        the second is the y-axis label.
    start_row
        First row for data insertion.

    Returns
    -------
    :
        The Excel bar chart object ready for insertion in a sheet.
    """
    nrows, ncols = df.shape

    # gapWidth controls the space between outer column level groups
    cat_len = len(df.columns.unique(1))
    chart_kwargs = {"gapWidth": 20} if cat_len > 1 else {}
    chart = BarChart(**chart_kwargs)

    # data includes index names left to numeric data
    min_col = df.index.nlevels
    min_row = df.columns.nlevels + 1 + start_row
    max_col = df.index.nlevels + ncols  # same
    max_row = df.columns.nlevels + start_row + nrows  # same

    if cfg.chart_switch_axis:
        # use column names as x-axis labels: the upper left cell shifts
        min_col += 1  # one to the right
        min_row -= 1  # one up

    data = Reference(ws, min_col, min_row, max_col, max_row)

    # reference is the horizontal header innermost one or two rows,
    # or the index in case switch_row_col is True
    min_col = df.index.nlevels + 1
    max_col = df.index.nlevels + ncols
    min_row = start_row + df.columns.nlevels - (1 if cat_len > 1 else 0)
    max_row = start_row + df.columns.nlevels

    if cfg.chart_switch_axis:
        # the categories become the index names of the column names
        # instead. The reference area transposes from row selection to
        # column selection.
        min_col = max_col = min_col - 1  # one left
        min_row += 2  # one down
        max_row = df.columns.nlevels + start_row + nrows

    cats = Reference(ws, min_col, min_row, max_col, max_row)

    chart.type = "col"
    chart.grouping = cfg.chart
    title = cfg.chart_title or df.columns.get_level_values("metric")[0]
    chart.title = title.format(location=ws.title, unit=axis_labels[1])

    # only stack chart if stacked otherwise 0 to avoid always
    # plotting a stacked chart (=100%)
    chart.overlap = 100 if cfg.chart == "stacked" else 0

    chart.add_data(data, from_rows=not cfg.chart_switch_axis, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.title = axis_labels[0]
    chart.y_axis.title = axis_labels[1]
    # excel row height is in pixel. resolution is 72 DPI.
    # default row height in Excel is 15
    height_factor = 0.53  # 15 (px) / 72 (px/Inch) * 2.54 (cm/Inch) = (cm)
    height_data = nrows + df.columns.nlevels
    chart.height = min(height_data * height_factor, 10)
    chart.width = cfg.chart_width  # cm

    if cfg.chart_switch_axis:  # the legend is redundant, as is the x-axis label
        chart.legend = None
        chart.x_axis.title = None

    # set bar colors in chart
    for i, carrier in enumerate(df.index):
        if not cfg.chart_switch_axis:
            color = cfg.chart_colors.get(carrier)
        else:
            color = cfg.chart_colors.get(df.columns.unique("carrier")[0])

        if not color:
            continue

        if not cfg.chart_switch_axis:
            chart.series[i].graphicalProperties.solidFill = color
            chart.series[i].graphicalProperties.line.solidFill = color
        else:
            for ser in chart.series:
                single_bar = DataPoint(idx=i)
                single_bar.graphicalProperties.solidFill = color
                # white borders to separate neighbors of same color
                single_bar.graphicalProperties.line.solidFill = "FFFFFF"
                ser.data_points.append(single_bar)

    return chart


def _expand_column_to_fit_content(ws: Worksheet, df: pd.DataFrame, col: int) -> None:
    """
    Expand cell columns to improve readability in Excel.

    The function expands the column to the larger value of its current
    column width and the largest string in the input data frame index.

    Parameters
    ----------
    ws
        The open work sheet instance.
    df
        The data added to the worksheet.
    col
        The index of the column that should become expanded.
    """
    xl_col = xl_col_to_name(col)
    series = df.index if col == 0 else df[df.columns[col - 1]]
    data_width = series.astype(str).str.len().max()
    existing_width = ws.column_dimensions[xl_col].width
    column_width = max(existing_width, data_width)
    ws.column_dimensions[xl_col].width = column_width
